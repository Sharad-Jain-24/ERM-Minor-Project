#importing libraries & modules
import os
import re 
import cv2
import time
import pyqrcode
import numpy as np
import tkinter as tk
import pyzbar.pyzbar as pyzbar
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
from openpyxl.styles import colors
from openpyxl.styles.colors import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill

#path for excel file
path = "./data/regdata.xlsx"               

#code to register & generate QR for participant
def QRP():
    #codde for GUI of QR generator
    def QRGen():
        global screen5
        screen5 = Toplevel(screen4)
        screen5.title("QR Generator")
        screen5.geometry("405x615")
        screen5.resizable(False, False)
        screen5.config(background="green")
        label = Label(screen5, text="Enter Name : ", bg="green")
        label.grid(row=0, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrName)
        screen5.entry.grid(row=0, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="Enter Phno : ", bg="green")
        label.grid(row=1, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrphno)
        screen5.entry.grid(row=1, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="Enter Email : ", bg="green")
        label.grid(row=2, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrmail)
        screen5.entry.grid(row=2, column=2, padx=5, pady=5, columnspan=2)    
        label = Label(screen5, text="1st Event Name : ", bg="green")
        label.grid(row=3, column=1, padx=5, pady=5)    
        screen5.entry = Entry(screen5, width=30, textvariable=qrevent1)
        screen5.entry.grid(row=3, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="2nd Event Name : ", bg="green")
        label.grid(row=4, column=1, padx=5, pady=5)    
        screen5.entry = Entry(screen5, width=30, textvariable=qrevent2)
        screen5.entry.grid(row=4, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="QR Code : ", bg="green")
        label.grid(row=5, column=1, padx=5, pady=5)
        button = Button(screen5, width=10, text="Generate", command=QRCodeGenerate)
        button.grid(row=5, column=2, padx=5, pady=5, columnspan=1)
        buton = Button(screen5, width=10, text="Clear", command=QRClear)
        buton.grid(row=5, column=3, padx=5, pady=5, columnspan=1)
        screen5.imageLabel = Label(screen5, background="green")
        screen5.imageLabel.grid(row=6, column=1, columnspan=3, padx=5, pady=5)
        image = Image.open("./resc/wait.jpg")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image
        button = Button(screen5, width=13, text="Registration Open", command=QROpen)
        button.grid(row=7, column=1, padx=5, pady=5, columnspan=1)
        buton = Button(screen5, width=13, text="Registration Closed", command=QRClose)
        buton.grid(row=7, column=3, padx=5, pady=5, columnspan=1)
        
    #code for clearing values of GUI fields        
    def QRClear():        
        qrName.set("")
        qrphno.set("")
        qrmail.set("")
        qrevent1.set("")
        qrevent2.set("")
        image = Image.open("./resc/done.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image
        
    #codde to reactivate all disabled GUI fields    
    def QROpen():
        print("OPEN")
        """
        reactivate all disabled fields & buttons
        """
        
    #code to disable all GUI fields
    def QRClose():
        print("Close")        
        """
        ask for CAPTCHA type verification, then disable 
        all fields and buttons except Registration Open,
        break databases according to event names
        """
        
    #code to generate QR with participant data    
    def QRCodeGenerate():    
        if (qrName.get() != '') and (qrphno.get() != '') and (qrmail.get() != '') and (qrevent1.get() != ''):
            if (len(qrphno.get()) == 10):
                regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                if(re.search(regex,qrmail.get())):  
                    content = qrName.get() + "-" + qrphno.get()
                    qrGenerate = pyqrcode.create(content)
                    qrCodePath = './data/'
                    qrCodeName = qrCodePath + qrphno.get() + ".png"
                    qrGenerate.png(qrCodeName, scale=10)
                    image = Image.open(qrCodeName)
                    image = image.resize((350, 350), Image.ANTIALIAS)
                    image = ImageTk.PhotoImage(image)
                    screen5.imageLabel.config(image=image)
                    screen5.imageLabel.photo = image
                    QRdatamgSQL()
                    QRdatamgXL()
                else:  
                    messagebox.showerror("ERROR", "Invalid Email ID")
            else:
                messagebox.showerror("ERROR", "Invalid Phone Number")
        else:
            messagebox.showerror("ERROR", "Fields Incomplete")
                
    #code add participant ddata to excel sheet
    def QRdatamgXL():
        wb = load_workbook(path)
        sheet = wb.active
        row = ((qrName.get(), qrphno.get(), qrmail.get(), qrevent1.get()))
        sheet.append(row)
        wb.save(path)

    #code add participant ddata to SQL
    def QRdatamgSQL():
        print("SHARAD")
        """
        ek new function bana to check if data 
        being entered is already in db.
        ek new function bana for SQL db creation,
        iss function se SQL db connection and
        add collected data to database.
        database ke table mei 5 columns rakh,
        4 for collected data, 1 for marking present.
        add if to check if email already in db,
        if yes give GUI prompt to confirms    
        """

    #code initializing varaibles of GUI
    qrName = StringVar()
    qrphno = StringVar()
    qrmail = StringVar()
    workbook = Workbook()
    qrevent1 = StringVar()
    qrevent2 = StringVar()    
    QRGen()

QRP()

