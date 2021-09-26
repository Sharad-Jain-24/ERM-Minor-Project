#importing libraries & modules
import os
import re 
import cv2
import time
import pyqrcode
import numpy as np
import tkinter as tk
import pyzbar.pyzbar as pyzbar
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
from openpyxl.styles import colors
from openpyxl.styles.colors import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill

#path for excel file
path = "./data/regdata.xlsx"               

#QR Scanner code
def QRScan():
    #start device camera
    cap = cv2.VideoCapture(0)
    cap.set(3,640)
    cap.set(4,480)
    time.sleep(2)

    #find content of QR
    def decode(im): 
        decodedObjects = pyzbar.decode(im)
        return decodedObjects

    font = cv2.FONT_HERSHEY_SIMPLEX

    #check if data of scanned QR in excel    
    def regbk(bar):
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                bar = bar.replace("'","")
                value = cell.value + "'"
                if(cell.value == bar):
                    print(cell.coordinate)
                    cell.fill = PatternFill(bgColor="00FF00", fill_type="solid")
                    cell.font = Font(color="00FF00")
                    wb.save(path)

    #check if data of scanned QR in SQL
    def regdb():
        print ("Sharad")
        """
        iss function se connect kar existing SQL db se,
        aur check kar if data being scanned db mei hai ya nahi,
        if present add present in present column
        and scanner wale hull ka color green
        if already marked present hull color red,
        GUI alert pop karva.
        """

    #3scan the QR 
    def app():
        decodedObject = ""
        while(cap.isOpened()):
            ret, frame = cap.read()
            im = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            decodedObjects = decode(im)
            for decodedObject in decodedObjects: 
                points = decodedObject.polygon
                if len(points) > 4 : 
                  hull = cv2.convexHull(np.array([point for point in points], dtype=np.float32))
                  hull = list(map(tuple, np.squeeze(hull)))
                else : 
                  hull = points;         
                n = len(hull)     
                for j in range(0,n):
                  cv2.line(frame, hull[j], hull[ (j+1) % n], (255,0,0), 3)
                x = decodedObject.rect.left
                y = decodedObject.rect.top
                cv2.putText(frame, str(decodedObject.data), (x, y), font, 1, (0,255,255), 2, cv2.LINE_AA)
                
            cv2.imshow('frame', frame)
            key = cv2.waitKey(1)
            if key & 0xFF == ord('q'):
                break
            elif key & 0xFF == ord('s'):
                cv2.imwrite('./scan/capt.png', frame)     
            
        barCode = str(decodedObject.data)            
        barC = barCode.split('-')
        bar = barC[1]
        regbk(bar)
        regdb()
                                     
    app()
    cap.release()
    cv2.destroyAllWindows()

QRScan()

