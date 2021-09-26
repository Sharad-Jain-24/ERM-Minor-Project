#importing libraries & modules
import os
import re
import cv2
import time
import random
import platform
import pyqrcode
import numpy as np
import tkinter as tk
import pyzbar.pyzbar as pyzbar
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
from openpyxl.styles import colors
from openpyxl.styles.colors import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill
try:
    import App.frontend_api as fi
except:
    import frontend_api as fi

# check OS
def chkos():
    oname = platform.system()
    global screen1geo, screen1_5geo, screen2geo, screen3geo, screen4geo, screen5geo, screen6geo, screen7geo
    if oname == "Windows":
        screen1geo = "430x300"
        screen1_5geo = "490x145"
        screen2geo = "500x310"
        screen3geo = "360x125"
        screen4geo = "250x258"
        #screen5geo = "675x525"
        screen5geo = "690x470"
        screen6geo = "390x190"
        screen7geo = ""

    elif oname == "Linux":
        screen1geo = "375x300"
        screen1_5geo = "390x145"
        screen2geo = "520x310"
        screen3geo = "320x125"
        screen4geo = "250x258"
        #screen5geo = "760x420"
        screen5geo = "770x460"
        screen6geo = "455x190"
        screen7geo = ""

    else:
        #print ("Invalid OS")
        screen1geo = "375x300"
        screen1_5geo = "390x145"
        screen2geo = "520x310"
        screen3geo = "320x125"
        screen4geo = "250x258"
        #screen5geo = "760x420"
        screen5geo = "760x460"
        screen6geo = "455x190"
        screen7geo = ""

#path for excel file
path = "./data/regdata.xlsx"

#QR Scanner code
def QRScan():
    #start device camera
    cap = cv2.VideoCapture(0)
    cap.set(3,640)
    cap.set(4,480)
    time.sleep(2)

    """
    screen7 = Toplevel(screen4)
    screen7.title("QR Scanner")
    screen7.geometry(screen7geo)
    screen7.resizable(False, False)
    screen7.config(background=gcolor)
    screen7.focus_force()
    icon = PhotoImage(file="./resc/qr-code-scan.png")
    screen7.iconphoto(False, icon)
    """

    #find content of QR
    def decode(im):
        decodedObjects = pyzbar.decode(im)
        return decodedObjects

    font = cv2.FONT_HERSHEY_SIMPLEX

    #check if data of scanned QR in excel
    def regbk(bar):
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                bar = bar.replace("'","")
                value = cell.value + "'"
                if(cell.value == bar):
                    print(cell.coordinate)
                    cell.fill = PatternFill(bgColor="00FF00", fill_type="solid")
                    cell.font = Font(color="00FF00")
                    wb.save(path)

    #check if data of scanned QR in SQL
    def regdb():
        print ("Sharad")
        """
        iss function se connect kar existing SQL db se,
        aur check kar if data being scanned db mei hai ya nahi,
        if present add present in present column
        and scanner wale hull ka color green
        if already marked present hull color red,
        GUI alert pop karva.
        """

    #scan the QR
    def app():
        decodedObject = ""
        while(cap.isOpened()):
            ret, frame = cap.read()
            im = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            decodedObjects = decode(im)
            for decodedObject in decodedObjects:
                points = decodedObject.polygon
                if len(points) > 4 :
                  hull = cv2.convexHull(np.array([point for point in points], dtype=np.float32))
                  hull = list(map(tuple, np.squeeze(hull)))
                else :
                  hull = points;
                n = len(hull)
                for j in range(0,n):
                  #while scanning
                  cv2.line(frame, hull[j], hull[ (j+1) % n], (255,0,0), 3)
                x = decodedObject.rect.left
                y = decodedObject.rect.top
                cv2.putText(frame, str(decodedObject.data), (x, y), font, 1, (0,255,255), 2, cv2.LINE_AA)

            cv2.imshow('frame', frame)
            key = cv2.waitKey(1)
            if key & 0xFF == ord('q'):
                cap.release()
                cv2.destroyAllWindows()
                break
            elif key & 0xFF == ord('s'):
                if (bar != "") or (bar is not None):
                    iname = "./scan/" + bar + ".png"
                else:
                    iname = "./scan/" + random.randint(1, 101) + ".png"
                cv2.imwrite(iname, frame)

        try:
            barCode = str(decodedObject.data)
            barC = barCode.split('-')
            bar = barC[1]
            regbk(bar)
            regdb()
        except:
            messagebox.showerror("ALERT", "No QR Detected")

    app()
    cap.release()
    cv2.destroyAllWindows()
    screen4.focus_force()

#code to register & generate QR for participant
def QRP():
    #codde for GUI of QR generator
    def QRGen():
        global screen5
        gcolor = "#161a2d"
        screen5 = Toplevel(screen4)
        screen5.title("QR Generator")
        screen5.geometry(screen5geo)
        screen5.resizable(False, False)
        screen5.config(background=gcolor)
        screen5.focus_force()
        icon = PhotoImage(file="./resc/laptop.png")
        screen5.iconphoto(False, icon)
        label = Label(screen5, text="Event Registration", bg=gcolor, font=("Times New Roman", 20, 'bold'))
        label.configure(foreground="white", anchor="center")
        label.grid(row=0, column=2, padx=5, pady=5, columnspan=4)
        label = Label(screen5, text="Enter all details or QR-ID of participant", bg=gcolor)
        label.configure(foreground="white")
        label.grid(row=1, column=1, padx=5, pady=10, columnspan=3)
        label = Label(screen5, text="Enter Name : ", bg=gcolor)
        label.configure(foreground="white")
        label.grid(row=2, column=1, padx=5, pady=10)
        screen5.entryname = Entry(screen5, width=30, textvariable=qrName)
        screen5.entryname.grid(row=2, column=2, padx=5, pady=10, columnspan=2)
        screen5.entryname.focus_set()
        label = Label(screen5, text="Enter Phno : ", bg=gcolor)
        label.configure(foreground="white")
        label.grid(row=3, column=1, padx=5, pady=10)
        screen5.entryphno = Entry(screen5, width=30, textvariable=qrphno)
        screen5.entryphno.grid(row=3, column=2, padx=5, pady=10, columnspan=2)
        label = Label(screen5, text="Enter Email : ", bg=gcolor)
        label.configure(foreground="white")
        label.grid(row=4, column=1, padx=5, pady=10)
        screen5.entrymail = Entry(screen5, width=30, textvariable=qrmail)
        screen5.entrymail.grid(row=4, column=2, padx=5, pady=10, columnspan=2)
        label = Label(screen5, text="QR ID : ", bg=gcolor)
        label.configure(foreground="white")
        label.grid(row=5, column=1, padx=5, pady=10)
        screen5.entry = Entry(screen5, width=15, textvariable=qrID)
        #screen5.entry.grid(row=5, column=2, padx=5, pady=10, columnspan=2)
        
        screen5.entry.grid(row=5, column=2, padx=10, pady=10, columnspan=1, sticky='w')
        sbtn = Button(screen5, width=8, text="Scanner", command=QRScan)
        sbtn.grid(row=5, column=3, padx=5, pady=10, sticky='e')
        
        ttk.Separator(screen5, orient=HORIZONTAL).grid(column=1, row=6, columnspan=3, sticky='ew')
        label = Label(screen5, text="1st Event Name : ", bg=gcolor)
        label.configure(foreground="white")
        label.grid(row=7, column=1, padx=5, pady=10)
        label = Label(screen5, text="2nd Event Name : ", bg=gcolor)
        label.configure(foreground="white")
        label.grid(row=8, column=1, padx=5, pady=10)
        screen5.entry1 = ttk.Combobox(screen5, width=27, textvariable=qrevent1)#, state="readonly")
        screen5.entry1.grid(row=7, column=2, padx=5, pady=10, columnspan=2)
        screen5.entry1['values'] = () #add data from db here
        screen5.entry1.current()
        screen5.entry2 = ttk.Combobox(screen5, width=27, textvariable=qrevent2)#, state="readonly")
        screen5.entry2.grid(row=8, column=2, padx=5, pady=10, columnspan=2)
        screen5.entry2['values'] = () #add data from db here
        screen5.entry2.current()
        label = Label(screen5, text="QR Code : ", bg=gcolor)
        label.configure(foreground="white")
        label.grid(row=9, column=1, padx=5, pady=10)
        button = Button(screen5, width=10, text="Generate", command=QRCodeGenerate)
        button.grid(row=9, column=2, padx=5, pady=10, columnspan=1)
        screen5.bind('<Return>', lambda event=None: button.invoke())
        buton = Button(screen5, width=10, text="Clear", command=QRClear)
        buton.grid(row=9, column=3, padx=5, pady=10, columnspan=1)
        screen5.bind("<Control-r>", lambda event=None: buton.invoke())
        screen5.imageLabel = Label(screen5, background=gcolor)
        screen5.imageLabel.grid(row=2, column=4, rowspan=9, columnspan=3, padx=(10,5), pady=10)
        image = Image.open("./resc/wait.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image

    #reload wait image
    def ld():
        image = Image.open("./resc/wait.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image

    # code for clearing values of GUI fields
    def QRClear():
        screen5.entryname.focus_set()
        qrName.set("")
        qrphno.set("")
        qrmail.set("")
        qrevent1.set("")
        qrevent2.set("")
        qrID.set("")
        image = Image.open("./resc/done.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image
        screen5.after(500, ld)

    #code to generate QR with participant data
    def QRCodeGenerate():
        if (qrName.get() != '') and (qrphno.get() != '') and (qrmail.get() != '') and (qrevent1.get() != ''):
            if (len(qrphno.get()) == 10):
                try:
                    phno = int(qrphno.get())
                    rege = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                    if(re.search(rege,qrmail.get())):
                        content = qrName.get() + "-" + qrphno.get()
                        qrGenerate = pyqrcode.create(content)
                        qrCodePath = './data/'
                        qrCodeName = qrCodePath + qrphno.get() + ".png"
                        qrGenerate.png(qrCodeName, scale=10)
                        image = Image.open(qrCodeName)
                        image = image.resize((350, 350), Image.ANTIALIAS)
                        image = ImageTk.PhotoImage(image)
                        screen5.imageLabel.config(image=image)
                        screen5.imageLabel.photo = image
                        QRdatamgSQL()
                        QRdatamgXL()
                    else:
                        messagebox.showerror("ALERT", "Invalid Email ID")
                        screen5.focus_force()
                        screen5.entrymail.focus_set()
                except:
                    messagebox.showerror("ALERT", "Phone: Not a Number")
                    screen5.focus_force()
                    screen5.entryphno.focus_set()
            else:
                messagebox.showerror("ALERT", "Invalid Phone Number")
                screen5.focus_force()
                screen5.entryphno.focus_set()
        elif (qrID.get() != '') and (qrevent1.get() != ''):
            print ("lol")
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")
            screen5.focus_force()

    #code add participant ddata to excel sheet
    def QRdatamgXL():
        wb = load_workbook(path)
        sheet = wb.active
        row = ((qrName.get(), qrphno.get(), qrmail.get(), qrevent1.get()))
        sheet.append(row)
        wb.save(path)

    #code add participant ddata to SQL
    def QRdatamgSQL():
        print("SHARAD")
        """
        ek new function bana to check if data
        being entered is already in db.
        ek new function bana for SQL db creation,
        iss function se SQL db connection and
        add collected data to database.
        database ke table mei 5 columns rakh,
        4 for collected data, 1 for marking present.
        add if to check if email already in db,
        if yes give GUI prompt to confirms
        """

    #code initializing varaibles of GUI
    qrName = StringVar()
    qrphno = StringVar()
    qrmail = StringVar()
    workbook = Workbook()
    qrevent1 = StringVar()
    qrevent2 = StringVar()
    qrID = StringVar()
    QRGen()

#code to maintain event list
"""
sharad please decide how you want this function to work
"""
def eventmgm():
    global screen6
    #clear fields
    def clrevent():
        adevent.focus_set()
        evename.set("")
        evedate.set("")
        evetime.set("")

    #add events to database
    def addevent():
        if (adevent.get() != "") and (adeveti.get() != "") and (adevedt.get() != ""):
            print("add event to SQL")
            #add entered event in db
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")

    #remove events from database
    def remevent():
        if (adevent.get() != "") and (adeveti.get() != "") and (adevedt.get() != ""):
            print("remove event from sql")
            #remove entered event from db
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")

    #GUI code for event manager
    evename = StringVar()
    evetime = StringVar()
    evedate = StringVar()
    screen6 = Toplevel(screen4)
    screen6.title("Event Manager")
    screen6.geometry(screen6geo)
    screen6.resizable(False, False)
    screen6.config(background="green")
    icon = PhotoImage(file="./resc/team-management.png")
    screen6.iconphoto(False, icon)
    screen6.focus_force()
    label = Label(screen6, text="Event Management", bg="green", font=("Times New Roman", 20, 'bold'))
    label.configure(foreground="white", anchor="center")
    label.grid(row=0, column=1, padx=(17,0), pady=(10,15), columnspan=4)
    lbl = Label(screen6, text="Event Name", bg="green")
    lbl.configure(foreground="white")
    lbl.grid(row=1, column=1, padx=(40,5), pady=5, columnspan=1)
    adevent = Entry(screen6, width='17', textvariable=evename)
    adevent.grid(row=1, column=2, padx=5, pady=5, columnspan=1)
    adevent.focus_set()
    lbl = Label(screen6, text="Event Date", bg="green")
    lbl.configure(foreground="white")
    lbl.grid(row=2, column=1, padx=(40,5), pady=5, columnspan=1)
    adeveti = Entry(screen6, width='17', textvariable=evedate)
    adeveti.grid(row=2, column=2, padx=5, pady=5, columnspan=1)
    lbl = Label(screen6, text="Event Time", bg="green")
    lbl.configure(foreground="white")
    lbl.grid(row=3, column=1, padx=(40,5), pady=5, columnspan=1)
    adevedt = Entry(screen6, width='17', textvariable=evetime)
    adevedt.grid(row=3, column=2, padx=5, pady=5, columnspan=1)
    bnt = Button(screen6, text="Clear", command=clrevent, width='13')
    bnt.grid(row=1, column=3, padx=5, pady=5, columnspan=2)
    nbt = Button(screen6, text="Add Event", command=addevent, width='13')
    nbt.grid(row=2, column=3, padx=5, pady=5, columnspan=2)
    tnb = Button(screen6, text="Remove Event", command=remevent, width='13')
    tnb.grid(row=3, column=3, padx=5, pady=5, columnspan=2)
    screen6.bind("<Control-a>", lambda event=None: nbt.invoke())
    screen6.bind("<Control-d>", lambda event=None: tnb.invoke())
    screen6.bind("<Control-r>", lambda event=None: bnt.invoke())

    #code to monitor app close event
    def on_closing():
        screen4.deiconify()
        screen6.destroy()
    screen6.protocol("WM_DELETE_WINDOW", on_closing)

#code to manage organizer/user tasks
def mgm_page():
    #GUI for organizer management
    screen3.withdraw()
    global screen4, background_label
    screen4 = Toplevel(screen3)
    screen4.title("Select")
    screen4.geometry(screen4geo)
    screen4.resizable(False, False)
    screen4.config(background=colr)
    icon = PhotoImage(file="./resc/process.png")
    screen4.iconphoto(False, icon)
    screen4.focus_force()
    btn = Button(screen4, width=15, borderwidth=0, text="QR Generator", command=QRP)
    btn.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
    btn.place(relx=0.5, rely=0.25, anchor=N)
    bnt = Button(screen4, width=15, borderwidth=0, text="QR Scanner", command=QRScan)
    bnt.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
    bnt.place(relx=0.5, rely=0.5, anchor=CENTER)
    tbn = Button(screen4, width=15, borderwidth=0, text="Event Management", command=eventmgm)
    tbn.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
    tbn.place(relx=0.5, rely=0.75, anchor=S)
    screen4.bind("<Control-g>", lambda event=None: btn.invoke())
    screen4.bind("<Control-s>", lambda event=None: bnt.invoke())
    screen4.bind("<Control-e>", lambda event=None: tbn.invoke())

    #code to monitor app close event
    def on_closing():
        screen1.destroy()
    screen4.protocol("WM_DELETE_WINDOW", on_closing)

"""
sharad please optimize organizer registration &
login technique as you see fit.
"""
#GUI & code for login & signup
def main_page():
    global username_verify, password_verify, username1, password1

    #code to clear login data fields after successful login
    def clrlogin():
        username_verify.set("")
        password_verify.set("")
        username_entry1.focus_set()
        mgm_page()

    #code to organizer management
    def register_user():
        #code to monitor screen1 close event
        def on_closing():
            #screen1_5.destroy()
            screen2.destroy()
            screen1.deiconify()
        screen2.protocol("WM_DELETE_WINDOW", on_closing)

        #user add success
        def disab():
            global screen1_5
            screen1_5 = Toplevel(screen1)
            screen1_5.title("Success")
            screen1_5.geometry(screen1_5geo)
            screen1_5.resizable(False, False)
            screen1_5.config(background="green")
            screen1_5.focus_force()

            #code to call login success screen
            def calllog():
                screen1_5.destroy()
                screen2.destroy()
                adminlogin()

            label = Label(screen1_5, text="", bg="green")
            label.grid(row=1, column=1)
            label = Label(screen1_5, text="Registeration Success", width='30', bg="green", font=("Times New Roman", 20, 'bold'))
            label.configure(foreground="white")
            label.grid(pady=5, row=2, column=1, columnspan=1)
            bttn = Button(screen1_5, text="OK", width="15", command=calllog)
            bttn.grid(pady=5, row=3, column=1, columnspan=1)
            screen1_5.bind('<Return>', lambda event=None: bttn.invoke())

        disab()
        #check if password is strong
        # if re.fullmatch(r'[A-Za-z0-9@#$%^&+=]{8,}', password.get()):
        #     username_info = username.get()
        #     password_info = password.get()
        #     file = open(username_info, "w")
        #     file.write(username_info+"\n")
        #     file.write(password_info)
        #     file.close()
        #     disab()
        # else:
        #     messagebox.showerror("ALERT", "Password not Strong")

    #GUI code for adding organizer
    def register():
        global screen2, labl, buutn, username, password, username_entry, password_entry, emailid, phno, rights, emailid_entry, phno_entry, perm_entry, regbtn
        screen2 = Toplevel(screen1)
        screen2.title("Register")
        screen2.geometry(screen2geo)
        screen2.resizable(False, False)
        screen2.config(background=colr)
        screen3.focus_force()
        screen1.withdraw()
        icon = PhotoImage(file="./resc/add.png")
        screen2.iconphoto(False, icon)
        username = StringVar()
        password = StringVar()
        emailid = StringVar()
        phno = StringVar()
        rights = StringVar()

        def valinp():
            if (username_entry.get() != "") and (emailid_entry.get() != "") and (phno_entry.get() != "") and (password_entry.get() != "") and (perm_entry.get() != "Select"):
                if (len(phno_entry.get()) == 10):
                    rege = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                    if(re.search(rege, emailid_entry.get())):
                        if re.fullmatch(r'[A-Za-z0-9@#$%^&+=]{8,}', password.get()):
                            # register_user()
                            perm = perm_entry.get()
                            perm = 2 if perm == "Admin" else 1
                            resp = fi.add_user(name=username_entry.get(), email_id=emailid_entry.get(),
                                               password=password_entry.get(), phone=int(phno_entry.get()), perm=perm)
                            if resp == 0:
                                messagebox.showerror("ALERT", "User email already exists")
                            else:
                                register_user()
                        else:
                            messagebox.showerror("ALERT", "Password not Strong")
                    else:
                        messagebox.showerror("ALERT", "Invalid Email")
                else:
                    messagebox.showerror("ALERT", "Invalid Phone Number")
            else:
                messagebox.showerror("ALERT", "Fields Incomplete")

        labl = Label(screen2, text="Please enter user information", width="30", bg=colr)
        labl.configure(foreground="white", font=("Times New Roman", 20, 'bold'))
        labl.grid(row=1, column=1, padx=5, pady=5, columnspan=2)
        labl = Label(screen2, text="Username", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        username_entry = Entry(screen2, textvariable=username)
        username_entry.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        username_entry.focus_set()
        labl = Label(screen2, text="Email ID", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=2, column=2, padx=5, pady=5, columnspan=1)
        emailid_entry = Entry(screen2, textvariable=emailid)
        emailid_entry.grid(row=3, column=2, padx=5, pady=5, columnspan=1)
        labl = Label(screen2, text="Phone Number", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=4, column=1, padx=5, pady=5, columnspan=1)
        phno_entry = Entry(screen2, textvariable=phno)
        phno_entry.grid(row=5, column=1, padx=5, pady=5, columnspan=1)
        labl = Label(screen2, text="Password", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=4, column=2, padx=5, pady=5, columnspan=1)
        password_entry = Entry(screen2, show="*", textvariable=password)
        password_entry.grid(row=5, column=2, padx=5, pady=5, columnspan=1)
        labl = Label(screen2, text="", width="30", bg=colr)
        labl.grid(row=6, column=1, padx=5, pady=5, columnspan=2)
        labl = Label(screen2, text="Permission : ", width='30', bg=colr)
        labl.configure(foreground="white")
        labl.grid(row=7, column=1, padx=5, pady=5, columnspan=1)
        perm_entry = ttk.Combobox(screen2, textvariable=rights, width="17", values=["Select", "Admin", "User"], state="readonly")
        perm_entry.current(0)
        perm_entry.grid(row=7, column=2, columnspan=1, pady=5)
        labl = Label(screen2, text="", bg=colr)
        labl.grid(row=8, column=1, columnspan=2)
        regbtn = Button(screen2, text="Sumbit", width='18', command=valinp)
        regbtn.grid(row=9, column=1, padx=5, pady=5, columnspan=2)
        screen2.bind('<Return>', lambda event=None: regbtn.invoke())

        #code to monitor screen2 close event
        def on_closing():
            screen3.deiconify()
            screen2.destroy()
        screen2.protocol("WM_DELETE_WINDOW", on_closing)

    #GUI if data is of admin
    def adminlogin():
        screen3.geometry(screen3geo)
        label = Label(screen3, text="Login Success", width='30', bg="green")
        label.configure(foreground="white", font=("Times New Roman", 16, 'bold'))
        label.grid(row=1, column=1, pady=5, columnspan=1)
        bttnn = Button(screen3, text="OK", width="15", command=clrlogin)
        bttnn.grid(row=2, column=1, pady=5, columnspan=1)
        bttn = Button(screen3, text="Add Organizer", width="15", command=register)
        bttn.grid(row=3, column=1, pady=5, columnspan=1)
        screen3.bind('<Return>', lambda event=None: bttnn.invoke())
        screen3.bind("<Control-a>", lambda event=None: bttn.invoke())

    #GUI if data is of user
    def userlogin():
        screen3.geometry(screen3geo)
        label = Label(screen3, text="", bg="green")
        label.grid(row=1, column=1, pady=5)
        label = Label(screen3, text="Login Success", width='30', bg="green")
        label.configure(foreground="white", font=("Times New Roman", 16, 'bold'))
        label.grid(row=2, column=1, pady=5)
        bttn = Button(screen3, text="OK", width="10", command=clrlogin)
        bttn.grid(row=3, column=1, pady=5)
        screen3.bind('<Return>', lambda event=None: bttn.invoke())

    #code for GUI & user details verification
    def login_verify():
        screen1.withdraw()
        global screen3
        screen3 = Toplevel(screen1)
        screen3.title("Info")
        screen3.geometry(screen3geo)
        screen3.resizable(False, False)
        screen3.config(background="green")
        screen3.focus_force()
        icon = PhotoImage(file="./resc/check.png")
        screen3.iconphoto(False, icon)

        #code to monitor screen3 close event
        def on_closing():
            clrlogin()
            screen1.deiconify()
            screen3.destroy()
        screen3.protocol("WM_DELETE_WINDOW", on_closing)

        #validate dara in input
        username1 = username_verify.get()
        password1 = password_verify.get()
        # list_of_dir = os.listdir()
        # if username1 in list_of_dir:
        #     file = open (username1, "r")
        #     verify = file.read().splitlines()
        #     if (password1 and "Admin") in verify:
        #         adminlogin()
        #     elif password1 in verify:
        #         userlogin()
        #     else:
        #         on_closing()
        #         messagebox.showerror("ALERT", "Invalid Password")
        # else :
        #     on_closing()
        #     messagebox.showerror("ALERT", "Invalid User")

        resp = fi.login(uid=username1, password=password1)
        if resp == 2:
            adminlogin()
        elif resp == 1:
            userlogin()
        elif resp == 0:
            on_closing()
            messagebox.showerror("ALERT", "Invalid User/password")
        else:
            on_closing()
            messagebox.showerror("ALERT", "Invalid User")

    #code for login GUI
    def login():
        global screen1, username_verify, password_verify, username_entry1
        screen1 = Tk()
        screen1.title("Login")
        screen1.geometry(screen1geo)
        screen1.config(background=colr)
        screen1.resizable(False, False)
        icon = PhotoImage(file="./resc/login.png")
        screen1.iconphoto(False, icon)
        username_verify = StringVar()
        password_verify = StringVar()
        label = Label(text="", bg=colr)
        label.grid(row=1, column=1)
        label = Label(text="Please Enter your Login \nInformation", width='30', bg=colr)
        label.configure(foreground="white", font=("Times New Roman", 18, 'bold'))
        label.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        label = Label(text="Username : ", width='30', bg=colr)
        label.configure(foreground="white")
        label.grid(row=4, column=1, padx=25, pady=5, columnspan=1)
        username_entry1 = Entry(width="21", textvariable=username_verify)
        username_entry1.grid(row=5, column=1, padx=5, pady=5, columnspan=1)
        username_entry1.focus_set()
        label = Label(text="Password : ", width='30', bg=colr)
        label.configure(foreground="white")
        label.grid(row=6, column=1, padx=5, pady=5, columnspan=1)
        password_entry1 = Entry(width='21', show="*", textvariable=password_verify)
        password_entry1.grid(row=7, column=1, padx=5, pady=5, columnspan=1)
        label = Label(text="", bg=colr)
        label.grid(row=8, column=1)
        btnn = Button(text="Login", width="18", command=login_verify)
        btnn.grid(row=9, column=1, padx=5, pady=5, columnspan=1)
        screen1.bind('<Return>', lambda event=None: btnn.invoke())

        #monitor app close
        def on_closing(event):
            sys.exit()
        #binding Escape key as shortcut
        screen1.bind('<Escape>', on_closing)

        screen1.mainloop()

    #calling the login module
    login()

#setting a main colour theme
global colr
colr = "#1c44a5"

#checking OS
chkos()

#start of program
main_page()

