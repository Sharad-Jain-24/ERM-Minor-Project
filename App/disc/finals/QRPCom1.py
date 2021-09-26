#importing libraries & modules
import os
import re 
import cv2
import time
import pyqrcode
import numpy as np
import tkinter as tk
import pyzbar.pyzbar as pyzbar
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
from openpyxl.styles import colors
from openpyxl.styles.colors import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill

#path for excel file
path = "./data/regdata.xlsx"               

#QR Scanner code
def QRScan():
    #start device camera
    cap = cv2.VideoCapture(0)
    cap.set(3,640)
    cap.set(4,480)
    time.sleep(2)

    #find content of QR
    def decode(im): 
        decodedObjects = pyzbar.decode(im)
        return decodedObjects

    font = cv2.FONT_HERSHEY_SIMPLEX

    #check if data of scanned QR in excel    
    def regbk(bar):
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                bar = bar.replace("'","")
                value = cell.value + "'"
                if(cell.value == bar):
                    print(cell.coordinate)
                    cell.fill = PatternFill(bgColor="00FF00", fill_type="solid")
                    cell.font = Font(color="00FF00")
                    wb.save(path)

    #check if data of scanned QR in SQL
    def regdb():
        print ("Sharad")
        """
        iss function se connect kar existing SQL db se,
        aur check kar if data being scanned db mei hai ya nahi,
        if present add present in present column
        and scanner wale hull ka color green
        if already marked present hull color red,
        GUI alert pop karva.
        """

    #3scan the QR 
    def app():
        decodedObject = ""
        while(cap.isOpened()):
            ret, frame = cap.read()
            im = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            decodedObjects = decode(im)
            for decodedObject in decodedObjects: 
                points = decodedObject.polygon
                if len(points) > 4 : 
                  hull = cv2.convexHull(np.array([point for point in points], dtype=np.float32))
                  hull = list(map(tuple, np.squeeze(hull)))
                else : 
                  hull = points;         
                n = len(hull)     
                for j in range(0,n):
                  cv2.line(frame, hull[j], hull[ (j+1) % n], (255,0,0), 3)
                x = decodedObject.rect.left
                y = decodedObject.rect.top
                cv2.putText(frame, str(decodedObject.data), (x, y), font, 1, (0,255,255), 2, cv2.LINE_AA)
                
            cv2.imshow('frame', frame)
            key = cv2.waitKey(1)
            if key & 0xFF == ord('q'):
                break
            elif key & 0xFF == ord('s'):
                cv2.imwrite('./scan/capt.png', frame)     
            
        barCode = str(decodedObject.data)            
        barC = barCode.split('-')
        bar = barC[1]
        regbk(bar)
        regdb()

    app()
    cap.release()
    cv2.destroyAllWindows()

#code to register & generate QR for participant
def QRP():
    #codde for GUI of QR generator
    def QRGen():
        global screen5
        screen5 = Toplevel(screen4)
        screen5.title("QR Generator")
        screen5.geometry("405x615")
        screen5.resizable(False, False)
        screen5.config(background="green")
        label = Label(screen5, text="Enter Name : ", bg="green")
        label.grid(row=0, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrName)
        screen5.entry.grid(row=0, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="Enter Phno : ", bg="green")
        label.grid(row=1, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrphno)
        screen5.entry.grid(row=1, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="Enter Email : ", bg="green")
        label.grid(row=2, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrmail)
        screen5.entry.grid(row=2, column=2, padx=5, pady=5, columnspan=2)    
        label = Label(screen5, text="1st Event Name : ", bg="green")
        label.grid(row=3, column=1, padx=5, pady=5)    
        screen5.entry = Entry(screen5, width=30, textvariable=qrevent1)
        screen5.entry.grid(row=3, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="2nd Event Name : ", bg="green")
        label.grid(row=4, column=1, padx=5, pady=5)    
        screen5.entry = Entry(screen5, width=30, textvariable=qrevent2)
        screen5.entry.grid(row=4, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="QR Code : ", bg="green")
        label.grid(row=5, column=1, padx=5, pady=5)
        button = Button(screen5, width=10, text="Generate", command=QRCodeGenerate)
        button.grid(row=5, column=2, padx=5, pady=5, columnspan=1)
        buton = Button(screen5, width=10, text="Clear", command=QRClear)
        buton.grid(row=5, column=3, padx=5, pady=5, columnspan=1)
        screen5.imageLabel = Label(screen5, background="green")
        screen5.imageLabel.grid(row=6, column=1, columnspan=3, padx=5, pady=5)
        image = Image.open("./resc/wait.jpg")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image
        
    #code for clearing values of GUI fields        
    def QRClear():        
        qrName.set("")
        qrphno.set("")
        qrmail.set("")
        qrevent1.set("")
        qrevent2.set("")
        image = Image.open("./resc/done.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image
                
    #code to generate QR with participant data    
    def QRCodeGenerate():    
        if (qrName.get() != '') and (qrphno.get() != '') and (qrmail.get() != '') and (qrevent1.get() != ''):
            if (len(qrphno.get()) == 10):
                regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                if(re.search(regex,qrmail.get())):  
                    content = qrName.get() + "-" + qrphno.get()
                    qrGenerate = pyqrcode.create(content)
                    qrCodePath = './data/'
                    qrCodeName = qrCodePath + qrphno.get() + ".png"
                    qrGenerate.png(qrCodeName, scale=10)
                    image = Image.open(qrCodeName)
                    image = image.resize((350, 350), Image.ANTIALIAS)
                    image = ImageTk.PhotoImage(image)
                    screen5.imageLabel.config(image=image)
                    screen5.imageLabel.photo = image
                    QRdatamgSQL()
                    QRdatamgXL()
                else:  
                    messagebox.showerror("ERROR", "Invalid Email ID")
            else:
                messagebox.showerror("ERROR", "Invalid Phone Number")
        else:
            messagebox.showerror("ERROR", "Fields Incomplete")
                
    #code add participant ddata to excel sheet
    def QRdatamgXL():
        wb = load_workbook(path)
        sheet = wb.active
        row = ((qrName.get(), qrphno.get(), qrmail.get(), qrevent1.get()))
        sheet.append(row)
        wb.save(path)

    #code add participant ddata to SQL
    def QRdatamgSQL():
        print("SHARAD")
        """
        ek new function bana to check if data 
        being entered is already in db.
        ek new function bana for SQL db creation,
        iss function se SQL db connection and
        add collected data to database.
        database ke table mei 5 columns rakh,
        4 for collected data, 1 for marking present.
        add if to check if email already in db,
        if yes give GUI prompt to confirms    
        """

    #code initializing varaibles of GUI
    qrName = StringVar()
    qrphno = StringVar()
    qrmail = StringVar()
    workbook = Workbook()
    qrevent1 = StringVar()
    qrevent2 = StringVar()
    QRGen()

#code to maintain event list
"""
sharad please decide how you want this function to work
"""
def eventmgm():
    global values, txt
    values = []

    #code to make changes to events
    def QREventlist():
        #code to add events
        def addevent():
            if (addeventname.get() != ""):
                if (addeventname.get() not in values):
                    values.append(addeventname.get())
                    txt.insert(tk.END, values)
                    print(values)
                else:
                    print("already present")
            else:
                print(values)
            addeventname.set("")                                

        #code to remove events
        def remevent():
            if (reeventname.get() != ""):
                if (reeventname.get() in values):
                    values.remove(reeventname.get())
                    txt.delete(reeventname.get())
                    txt.update()
                    print(values)
                else:
                    print("not present")
            else:
                print(values)
            reeventname.set("")              
                    
        #code to clear GUI fields                    
        def clrevent():
            addeventname.set("")    
            reeventname.set("")              

        #code to create event management GUI
        screen4.withdraw()
        global screen5
        screen5 = Toplevel(screen4)
        screen5.title("Event List")
        screen5.geometry("340x160")
        screen5.config(background="green") 
        label = Label(screen5, text="Existing Events", width='17', bg='green')
        label.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
        
        """
        #view event names as drop down
        cmbx = ttk.Combobox(screen5, width='17', textvariable=eventval, state="readonly")
        cmbx['values'] = values
        cmbx.grid(row=1, column=2, padx=5, pady=5, columnspan=1)
        """
        
        #view event names as list
        txt = Text(screen5, width=20, height=1, state="readonly")
        txt.grid(row=1, column=2, padx=5, pady=5, columnspan=1)
        
        adevent = Entry(screen5, width='17', textvariable=addeventname)
        adevent.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        btn = Button(screen5, width=13, text="Add Event", command=addevent)
        btn.grid(row=2, column=2, padx=5, pady=5, columnspan=1)
        reevent = Entry(screen5, width='17', textvariable=reeventname)
        reevent.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        bnt = Button(screen5, width=13, text="Remove Event", command=remevent)
        bnt.grid(row=3, column=2, padx=5, pady=5, columnspan=1)
        nbt = Button(screen5, width=26, text="Clear", command=clrevent)
        nbt.grid(row=4, column=1, padx=5, pady=5, columnspan=2)        

        #code to monitor app close event
        def on_closing(event):
            screen4.deiconify()    
            screen5.destroy()
        screen5.protocol("WM_DELETE_WINDOW", on_closing)
        screen5.bind('<Escape>', on_closing)

    #code to initialze GUI varaiables
    eventval = StringVar()
    addeventname = StringVar()
    addevent = StringVar()
    reeventname = StringVar()
    remevent = StringVar()
    QREventlist()        
    
#code to make organizer taskss
def mgm_page():
    #GUI for organizer management
    screen3.withdraw()
    global screen4
    screen4 = Toplevel(screen3)
    screen4.title("Select")        
    #screen4.geometry("140x128")
    screen4.geometry("250x258")
    screen4.resizable(False, False)
    screen4.config(background="#856ff8")    
    btn = Button(screen4, width=13, text="QR Generator", command=QRP)
    btn.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
    bnt = Button(screen4, width=13, text="QR Scanner", command=QRScan)
    bnt.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
    tbn = Button(screen4, width=13, text="Event Management", command=eventmgm)
    tbn.grid(row=3, column=1, padx=5, pady=5, columnspan=1)

    #code to monitor app close event
    def on_closing(event):
        screen2.destroy()
    screen4.protocol("WM_DELETE_WINDOW", on_closing)
    screen4.bind('<Escape>', on_closing)
    
def main_page():
    global username_verify, password_verify, username1, password1        
    
    #code for organizer details verification
    def login_verify():
        screen2.withdraw()
        global screen3        
        screen3 = Toplevel(screen2)
        screen3.geometry("150x30")        

        #code to monitor app close event
        def on_closing(event):
            clrlogin()
            screen2.deiconify()    
            screen3.destroy()        
        screen3.protocol("WM_DELETE_WINDOW", on_closing)
        screen3.bind('<Escape>', on_closing)
        
        username1 = username_verify.get()
        password1 = password_verify.get()
        list_of_dir = os.listdir()
        if username1 in list_of_dir:
            file = open (username1, "r")
            verify = file.read().splitlines()
            if (password1 and "Admin") in verify:
                screen3.title("Info")
                screen3.geometry("250x115")
                label = Label(screen3, text="Login Success", width='30')
                label.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
                bttn = Button(screen3, text="OK", width="10", command=clrlogin)
                bttn.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
                bttn = Button(screen3, text="Add Organizer", width="10", command=register)
                bttn.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
            elif password1 in verify:
                screen3.title("Info")
                screen3.geometry("250x75")
                label = Label(screen3, text="Login Success", width='30')
                label.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
                bttn = Button(screen3, text="OK", width="10", command=clrlogin)
                bttn.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
            else:
                screen3.title("Alert")
                label = Label(screen3, text="Incorrect Password", width='17')
                label.grid(row=1, column=0, columnspan=1)                
        else :
            screen3.title("Alert")
            label = Label(screen3, text="Invalid User", width='17')
            label.grid(row=1, column=0, columnspan=1)
    
    #code to organizer management
    def register_user():
        #code to monitor app close event
        def on_closing(event):
            screen2.deiconify()    
            screen1.destroy()
        screen1.protocol("WM_DELETE_WINDOW", on_closing)
        screen1.bind('<Escape>', on_closing)
        
        #code to disable GUI fields for organizer management
        def disab():
            screen1.geometry("300x80")
            labl.grid_forget()
            username_entry.grid_forget()
            password_entry.grid_forget()
            buutn.grid_forget()
            
        def reab():
            register()                   

        #code for GUI of organizer registration success
        """
        sharad please optimize organizer registration & 
        login technique as you see fit.
        """

        #check if password is strong
        if re.fullmatch(r'[A-Za-z0-9@#$%^&+=]{8,}', password.get()):
            username_info = username.get()
            password_info = password.get()
            file = open(username_info, "w")
            file.write(username_info+"\n")
            file.write(password_info)
            file.close()
            disab()
            label = Label(screen1, text="Registeration Success", font ="green", width='30')
            label.grid(row=1, column=1, columnspan=1)
            bttn = Button(screen1, text="OK", width="15", command=on_closing)
            bttn.grid(row=2, column=1, columnspan=1)
        else:
            disab()
            label = Label(screen1, text="Use Strong Password", font ="green", width='30')
            label.grid(row=1, column=1, columnspan=1)
            bttn = Button(screen1, text="OK", width="15", command=reab)
            bttn.grid(row=2, column=1, columnspan=1)

    #code for organizer adding GUI
    def register():
        global screen1
        screen1 = Toplevel(screen2)
        screen1.geometry("250x400")
        screen1.title("Register")
        screen2.withdraw()
        screen3.withdraw()
        global labl
        global buutn
        global username
        global password
        global username_entry
        global password_entry
        username = StringVar()
        password = StringVar()
        emailid = StringVar()
        phno = StringVar() 
        rights =  StringVar()
        labl = Label(screen1, text="Please enter user information", width="30")
        labl.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
        labl = Label(screen1, text="Username", width='30')
        labl.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        username_entry = Entry(screen1, textvariable=username)
        username_entry.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        username_entry.focus_set()
        labl = Label(screen1, text="Email ID", width='30')
        labl.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        emailid_entry = Entry(screen1, textvariable=emailid)
        emailid_entry.grid(row=4, column=1, padx=5, pady=5, columnspan=1)
        
        labl = Label(screen1, text="Password", width='30')
        labl.grid(row=4, column=1, padx=5, pady=5, columnspan=1)
        password_entry = Entry(screen1, show="*", textvariable=password)
        password_entry.grid(row=5, column=1, padx=5, pady=5, columnspan=1)
        buutn = Button(screen1, text="Sumbit", width='18', command=register_user)
        buutn.grid(row=6, column=1, padx=5, pady=5, columnspan=1)
        screen1.bind('<Return>', lambda event=None: buutn.invoke())

        #code to monitor app close event
        def on_closing(event):
            clrlogin()
            screen2.deiconify()    
            screen1.destroy()
        screen1.protocol("WM_DELETE_WINDOW", on_closing)
        screen1.bind('<Escape>', on_closing)

    #code to clear login data fields after successful login
    def clrlogin():
        username_entry1.focus_set()
        username_verify.set("")
        password_verify.set("")
        mgm_page()

    #code for organizer login GUI
    def login():
        global screen2, username_verify, password_verify, username_entry1
        screen2 = Tk()
        screen2.geometry("255x200")
        screen2.title("Login")
        username_verify = StringVar()
        password_verify = StringVar()    
        label = Label(text="Please enter your login information", width='30')
        label.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
        label = Label(text="Username : ", width='30')
        label.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        username_entry1 = Entry(width="30", textvariable=username_verify)
        username_entry1.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        username_entry1.focus_set()        
        label = Label(text="Password : ", width='30')
        label.grid(row=4, column=1, padx=5, pady=5, columnspan=1)
        password_entry1 = Entry(width='30', show="*", textvariable=password_verify)
        password_entry1.grid(row=5, column=1, padx=5, pady=5, columnspan=1)
        btnn = Button(text="Login", width="18", command=login_verify)
        btnn.grid(row=6, column=1, padx=5, pady=5, columnspan=1)        
        screen2.bind('<Return>', lambda event=None: btnn.invoke())
        
        #monitor app close
        def on_closing(event):
            sys.exit()
        screen2.bind('<Escape>', on_closing)
        
        screen2.mainloop()

    login()

main_page()

