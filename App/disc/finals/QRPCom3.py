#importing libraries & modules
import os
import re 
import cv2
import time
import random
import pyqrcode
import numpy as np
import tkinter as tk
import pyzbar.pyzbar as pyzbar
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
from openpyxl.styles import colors
from openpyxl.styles.colors import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill

#path for excel file
path = "./data/regdata.xlsx"               

#QR Scanner code
def QRScan():
    #start device camera
    cap = cv2.VideoCapture(0)
    cap.set(3,640)
    cap.set(4,480)
    time.sleep(2)
    
    #find content of QR
    def decode(im): 
        decodedObjects = pyzbar.decode(im)
        return decodedObjects

    font = cv2.FONT_HERSHEY_SIMPLEX

    #check if data of scanned QR in excel    
    def regbk(bar):
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                bar = bar.replace("'","")
                value = cell.value + "'"
                if(cell.value == bar):
                    print(cell.coordinate)
                    cell.fill = PatternFill(bgColor="00FF00", fill_type="solid")
                    cell.font = Font(color="00FF00")
                    wb.save(path)

    #check if data of scanned QR in SQL
    def regdb():
        print ("Sharad")
        """
        iss function se connect kar existing SQL db se,
        aur check kar if data being scanned db mei hai ya nahi,
        if present add present in present column
        and scanner wale hull ka color green
        if already marked present hull color red,
        GUI alert pop karva.
        """

    #scan the QR 
    def app():
        decodedObject = ""
        while(cap.isOpened()):
            ret, frame = cap.read()
            im = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            decodedObjects = decode(im)
            for decodedObject in decodedObjects: 
                points = decodedObject.polygon
                if len(points) > 4 : 
                  hull = cv2.convexHull(np.array([point for point in points], dtype=np.float32))
                  hull = list(map(tuple, np.squeeze(hull)))
                else : 
                  hull = points;         
                n = len(hull)     
                for j in range(0,n):
                  cv2.line(frame, hull[j], hull[ (j+1) % n], (255,0,0), 3)
                x = decodedObject.rect.left
                y = decodedObject.rect.top
                cv2.putText(frame, str(decodedObject.data), (x, y), font, 1, (0,255,255), 2, cv2.LINE_AA)
                
            cv2.imshow('frame', frame)
            key = cv2.waitKey(1)
            if key & 0xFF == ord('q'):
                cap.release()
                cv2.destroyAllWindows()
                break
            elif key & 0xFF == ord('s'):
                if (bar != "") or (bar is not None):
                    iname = "./scan/" + bar + ".png"
                else:
                    iname = "./scan/" + random.randint(1, 101) + ".png"
                cv2.imwrite(iname, frame)     
            
        barCode = str(decodedObject.data)            
        barC = barCode.split('-')
        bar = barC[1]
        regbk(bar)
        regdb()

    app()
    cap.release()
    cv2.destroyAllWindows()

#code to register & generate QR for participant
def QRP():
    #codde for GUI of QR generator
    def QRGen():
        global screen5
        screen5 = Toplevel(screen4)
        screen5.title("QR Generator")
        screen5.geometry("405x570")
        screen5.resizable(False, False)
        screen5.config(background="green")
        label = Label(screen5, text="Enter Name : ", bg="green")
        label.grid(row=0, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrName)
        screen5.entry.grid(row=0, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="Enter Phno : ", bg="green")
        label.grid(row=1, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrphno)
        screen5.entry.grid(row=1, column=2, padx=5, pady=5, columnspan=2)
        label = Label(screen5, text="Enter Email : ", bg="green")
        label.grid(row=2, column=1, padx=5, pady=5)
        screen5.entry = Entry(screen5, width=30, textvariable=qrmail)
        screen5.entry.grid(row=2, column=2, padx=5, pady=5, columnspan=2)    
        label = Label(screen5, text="1st Event Name : ", bg="green")
        label.grid(row=3, column=1, padx=5, pady=5)    
        label = Label(screen5, text="2nd Event Name : ", bg="green")
        label.grid(row=4, column=1, padx=5, pady=5)
        screen5.entry1 = ttk.Combobox(screen5, width=29, textvariable=qrevent1)#, state="readonly")
        screen5.entry1.grid(row=3, column=2, padx=5, pady=5, columnspan=2)
        screen5.entry1['values'] = () #add data from db here
        screen5.entry1.current() 
        screen5.entry2 = ttk.Combobox(screen5, width=29, textvariable=qrevent2)#, state="readonly")
        screen5.entry2.grid(row=4, column=2, padx=5, pady=5, columnspan=2)
        screen5.entry2['values'] = () #add data from db here
        screen5.entry2.current() 
        label = Label(screen5, text="QR Code : ", bg="green")
        label.grid(row=5, column=1, padx=5, pady=5)
        button = Button(screen5, width=10, text="Generate", command=QRCodeGenerate)
        button.grid(row=5, column=2, padx=5, pady=5, columnspan=1)
        buton = Button(screen5, width=10, text="Clear", command=QRClear)
        buton.grid(row=5, column=3, padx=5, pady=5, columnspan=1)
        screen5.imageLabel = Label(screen5, background="green")
        screen5.imageLabel.grid(row=6, column=1, columnspan=3, padx=5, pady=5)
        image = Image.open("./resc/wait.jpg")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image
        
    #code for clearing values of GUI fields        
    def QRClear():        
        qrName.set("")
        qrphno.set("")
        qrmail.set("")
        qrevent1.set("")
        qrevent2.set("")
        image = Image.open("./resc/done.png")
        image = image.resize((350, 350), Image.ANTIALIAS)
        image = ImageTk.PhotoImage(image)
        screen5.imageLabel.config(image=image)
        screen5.imageLabel.photo = image
                
    #code to generate QR with participant data    
    def QRCodeGenerate():    
        if (qrName.get() != '') and (qrphno.get() != '') and (qrmail.get() != '') and (qrevent1.get() != ''):
            if (len(qrphno.get()) == 10):
                rege = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
                if(re.search(rege,qrmail.get())):  
                    content = qrName.get() + "-" + qrphno.get()
                    qrGenerate = pyqrcode.create(content)
                    qrCodePath = './data/'
                    qrCodeName = qrCodePath + qrphno.get() + ".png"
                    qrGenerate.png(qrCodeName, scale=10)
                    image = Image.open(qrCodeName)
                    image = image.resize((350, 350), Image.ANTIALIAS)
                    image = ImageTk.PhotoImage(image)
                    screen5.imageLabel.config(image=image)
                    screen5.imageLabel.photo = image
                    QRdatamgSQL()
                    QRdatamgXL()
                else:  
                    messagebox.showerror("ALERT", "Invalid Email ID")
            else:
                messagebox.showerror("ALERT", "Invalid Phone Number")
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")
                
    #code add participant ddata to excel sheet
    def QRdatamgXL():
        wb = load_workbook(path)
        sheet = wb.active
        row = ((qrName.get(), qrphno.get(), qrmail.get(), qrevent1.get()))
        sheet.append(row)
        wb.save(path)

    #code add participant ddata to SQL
    def QRdatamgSQL():
        print("SHARAD")
        """
        ek new function bana to check if data 
        being entered is already in db.
        ek new function bana for SQL db creation,
        iss function se SQL db connection and
        add collected data to database.
        database ke table mei 5 columns rakh,
        4 for collected data, 1 for marking present.
        add if to check if email already in db,
        if yes give GUI prompt to confirms    
        """

    #code initializing varaibles of GUI
    qrName = StringVar()
    qrphno = StringVar()
    qrmail = StringVar()
    workbook = Workbook()
    qrevent1 = StringVar()
    qrevent2 = StringVar()
    QRGen()

#code to maintain event list
"""
sharad please decide how you want this function to work
"""
def eventmgm():
    global screen5, evename, evetime, evedate, adevent, adevedt, adeveti
           
    #code to create event management GUI
    def evemg():
        #clear fields
        def clrevent():
            evename.set("")
            evedate.set("")
            evetime.set("")
            
        #add events to database    
        def addevent():
            print("HI")
            """
            add entered event in db
            """
            
        #remove events from database    
        def remevent():
            print("HI")
            """
            remove entered event frokm db
            """

        screen4.withdraw()
        evename = StringVar()
        evetime = StringVar()
        evedate = StringVar()
        screen5 = Toplevel(screen4)
        screen5.title("Event Manager")
        screen5.geometry("370x130")
        screen5.config(background="green") 
        lbl = Label(screen5, text="Event Name", bg="green")
        lbl.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
        adevent = Entry(screen5, width='17', textvariable=evename)
        adevent.grid(row=1, column=2, padx=5, pady=5, columnspan=1)       
        lbl = Label(screen5, text="Event Date", bg="green")
        lbl.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        adeveti = Entry(screen5, width='17', textvariable=evedate)
        adeveti.grid(row=2, column=2, padx=5, pady=5, columnspan=1)        
        lbl = Label(screen5, text="Event Time", bg="green")
        lbl.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        adevedt = Entry(screen5, width='17', textvariable=evetime)
        adevedt.grid(row=3, column=2, padx=5, pady=5, columnspan=1)                        
        nbt = Button(screen5, text="Clear", command=clrevent, width='10')
        nbt.grid(row=1, column=3, padx=5, pady=5, columnspan=2)        
        nbt = Button(screen5, text="Add Event", command=addevent, width='10')
        nbt.grid(row=2, column=3, padx=5, pady=5, columnspan=2)        
        nbt = Button(screen5, text="Remove Event", command=remevent, width='10')
        nbt.grid(row=3, column=3, padx=5, pady=5, columnspan=2)        

        #code to monitor app close event
        def on_closing():
            screen4.deiconify()    
            screen5.destroy()
        screen5.protocol("WM_DELETE_WINDOW", on_closing)

    evemg()        
    
#code to make organizer taskss
def mgm_page():
    #GUI for organizer management
    screen3.withdraw()
    global screen4
    screen4 = Toplevel(screen3)
    screen4.title("Select")      
    screen4.geometry("250x258")
    screen4.resizable(False, False)
    screen4.config(background="#856ff8")    
    btn = Button(screen4, width=13, text="QR Generator", command=QRP)
    btn.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
    btn.place(relx=0.5, rely=0.25, anchor=N)
    bnt = Button(screen4, width=13, text="QR Scanner", command=QRScan)
    bnt.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
    bnt.place(relx=0.5, rely=0.5, anchor=CENTER)
    tbn = Button(screen4, width=13, text="Event Management", command=eventmgm)
    tbn.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
    tbn.place(relx=0.5, rely=0.75, anchor=S)

    #code to monitor app close event
    def on_closing():
        screen2.destroy()
    screen4.protocol("WM_DELETE_WINDOW", on_closing)
    
"""
sharad please optimize organizer registration & 
login technique as you see fit.
"""   
#GUI & code for login & signup
def main_page():
    global username_verify, password_verify, username1, password1            
    
    #code to clear login data fields after successful login
    def clrlogin():
        username_entry1.focus_set()
        username_verify.set("")
        password_verify.set("")
        mgm_page()
        
    #code to organizer management
    def register_user():
    
        #code to monitor screen1 close event
        def on_closing():    
            screen1.destroy()
            screen2.deiconify()
        screen1.protocol("WM_DELETE_WINDOW", on_closing)
        
        #user add success
        def disab():
            global screen1_5
            screen1_5 = Toplevel(screen1)
 
            #code to call login success screen
            def calllog():    
                screen1_5.destroy()
                screen1.destroy()
                adminlogin()

            label = Label(screen1_5, text="Registeration Success", font ="green", width='30')
            label.grid(row=1, column=1, columnspan=1)
            bttn = Button(screen1_5, text="OK", width="15", command=calllog)
            bttn.grid(row=2, column=1, columnspan=1)
                                
        #check if password is strong
        if re.fullmatch(r'[A-Za-z0-9@#$%^&+=]{8,}', password.get()):
            username_info = username.get()
            password_info = password.get()
            file = open(username_info, "w")
            file.write(username_info+"\n")
            file.write(password_info)
            file.close()
            disab()
        else:
            messagebox.showerror("ALERT", "Password not Strong")

    #GUI code for adding organizer
    def register():
        global screen1
        screen1 = Toplevel(screen2)
        screen1.geometry("250x430")
        screen1.title("Register")
        screen2.withdraw()
        global labl, buutn, username, password, username_entry, password_entry, emailid, phno, rights, emailid_entry, phno_entry, opt1_entry, opt2_entry
        username = StringVar()
        password = StringVar()
        emailid = StringVar()
        phno = StringVar() 
        rights =  IntVar()
        labl = Label(screen1, text="Please enter user information", width="30")
        labl.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
        labl = Label(screen1, text="Username", width='30')
        labl.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        username_entry = Entry(screen1, textvariable=username)
        username_entry.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        username_entry.focus_set()
        labl = Label(screen1, text="Email ID", width='30')
        labl.grid(row=4, column=1, padx=5, pady=5, columnspan=1)
        emailid_entry = Entry(screen1, textvariable=emailid)
        emailid_entry.grid(row=5, column=1, padx=5, pady=5, columnspan=1)
        labl = Label(screen1, text="Phone Number", width='30')
        labl.grid(row=6, column=1, padx=5, pady=5, columnspan=1)
        phno_entry = Entry(screen1, textvariable=phno)
        phno_entry.grid(row=7, column=1, padx=5, pady=5, columnspan=1)
        labl = Label(screen1, text="Rights", width='30')
        labl.grid(row=8, column=1, padx=5, pady=5, columnspan=1)        
        opt1_entry = Radiobutton(screen1, variable=rights, value="Admin", width="10", state=NORMAL, anchor="w", text="Admin")
        opt1_entry.grid(row=9, column=1, padx=5, pady=5, columnspan=1)
        opt2_entry = Radiobutton(screen1, variable=rights, value="User", width="10", state=NORMAL, anchor="w", text="User")
        opt2_entry.grid(row=10, column=1, padx=5, pady=5, columnspan=1)     
        labl = Label(screen1, text="Password", width='30')
        labl.grid(row=11, column=1, padx=5, pady=5, columnspan=1)
        password_entry = Entry(screen1, show="*", textvariable=password)
        password_entry.grid(row=12, column=1, padx=5, pady=5, columnspan=1)
        buutn = Button(screen1, text="Sumbit", width='18', command=register_user)
        buutn.grid(row=13, column=1, padx=5, pady=5, columnspan=1)
        
        #check for all fields completion
        if (username != "") and (password != "") and (emailid != "") and (phno != "") and (rights != ""):
            screen1.bind('<Return>', lambda event=None: buutn.invoke())
        else:
            messagebox.showerror("ALERT", "Fields Incomplete")

        #code to monitor screen1 close event
        def on_closing():
            screen3.deiconify()
            screen1.destroy()
        screen1.protocol("WM_DELETE_WINDOW", on_closing)

    #GUI if data of admin
    def adminlogin():
        screen3.title("Info")
        screen3.geometry("250x115")
        label = Label(screen3, text="Login Success", width='30')
        label.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
        bttn = Button(screen3, text="OK", width="10", command=clrlogin)
        bttn.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        bttn = Button(screen3, text="Add Organizer", width="10", command=register)
        bttn.grid(row=3, column=1, padx=5, pady=5, columnspan=1)        
        
    #GUI if data of user
    def userlogin():        
        screen3.title("Info")
        screen3.geometry("250x75")
        label = Label(screen3, text="Login Success", width='30')
        label.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
        bttn = Button(screen3, text="OK", width="10", command=clrlogin)
        bttn.grid(row=2, column=1, padx=5, pady=5, columnspan=1)

    #code for organizer details verification
    def login_verify():
        screen2.withdraw()
        global screen3        
        screen3 = Toplevel(screen2)
        screen3.geometry("150x30")        

        #code to monitor screen3 close event
        def on_closing():
            clrlogin()
            screen2.deiconify()    
            screen3.destroy()        
        screen3.protocol("WM_DELETE_WINDOW", on_closing)
                
        #validate input        
        username1 = username_verify.get()
        password1 = password_verify.get()
        list_of_dir = os.listdir()
        if username1 in list_of_dir:
            file = open (username1, "r")
            verify = file.read().splitlines()
            if (password1 and "Admin") in verify:
                adminlogin()
            elif password1 in verify:
                userlogin()
            else:
                on_closing()
                messagebox.showerror("ALERT", "Invalid Password")
        else :
            on_closing()
            messagebox.showerror("ALERT", "Invalid User")

    #code for login GUI
    def login():
        global screen2, username_verify, password_verify, username_entry1
        screen2 = Tk()
        screen2.geometry("255x200")
        screen2.title("Login")
        username_verify = StringVar()
        password_verify = StringVar()    
        label = Label(text="Please enter your login information", width='30')
        label.grid(row=1, column=1, padx=5, pady=5, columnspan=1)
        label = Label(text="Username : ", width='30')
        label.grid(row=2, column=1, padx=5, pady=5, columnspan=1)
        username_entry1 = Entry(width="30", textvariable=username_verify)
        username_entry1.grid(row=3, column=1, padx=5, pady=5, columnspan=1)
        username_entry1.focus_set()        
        label = Label(text="Password : ", width='30')
        label.grid(row=4, column=1, padx=5, pady=5, columnspan=1)
        password_entry1 = Entry(width='30', show="*", textvariable=password_verify)
        password_entry1.grid(row=5, column=1, padx=5, pady=5, columnspan=1)
        btnn = Button(text="Login", width="18", command=login_verify)
        btnn.grid(row=6, column=1, padx=5, pady=5, columnspan=1)        
        screen2.bind('<Return>', lambda event=None: btnn.invoke())

        #monitor app close
        def on_closing(event):
            sys.exit()
        screen2.bind('<Escape>', on_closing)

        screen2.mainloop()

    login()

main_page()

