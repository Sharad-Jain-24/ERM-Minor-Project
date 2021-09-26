import re 
import pyqrcode
import tkinter as tk
from tkinter import *
from PIL import ImageTk, Image
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

def CreateWidgets():
    label = Label(text="Enter Name : ", bg="green")
    label.grid(row=0, column=1, padx=5, pady=5)
    root.entry = Entry(width=30, textvariable=qrInput)
    root.entry.grid(row=0, column=2, padx=5, pady=5, columnspan=2)
    label = Label(text="Enter Phno : ", bg="green")
    label.grid(row=1, column=1, padx=5, pady=5)
    root.entry = Entry(width=30, textvariable=qrphno)
    root.entry.grid(row=1, column=2, padx=5, pady=5, columnspan=2)
    label = Label(text="Enter Email : ", bg="green")
    label.grid(row=2, column=1, padx=5, pady=5)
    root.entry = Entry(width=30, textvariable=qrmail)
    root.entry.grid(row=2, column=2, padx=5, pady=5, columnspan=2)    
    label = Label(text="Event Name : ", bg="green")
    label.grid(row=3, column=1, padx=5, pady=5)    
    root.entry = Entry(width=30, textvariable=qrevent)
    root.entry.grid(row=3, column=2, padx=5, pady=5, columnspan=2)
    label = Label(text="QR Code : ", bg="green")
    label.grid(row=4, column=1, padx=5, pady=5)
    button = Button(width=10, text="Generate", command=QRCodeGenerate)
    button.grid(row=4, column=2, padx=5, pady=5, columnspan=1)
    buton = Button(width=10, text="Clear", command=QRClear)
    buton.grid(row=4, column=3, padx=5, pady=5, columnspan=1)
    label = Label(text="QR Code : ", bg="green")
    label.grid(row=4, column=1, padx=5, pady=5)
    root.imageLabel = Label(root, background="green")
    root.imageLabel.grid(row=5, column=1, columnspan=3, padx=5, pady=5)
    image = Image.open("./resc/wait.jpg")
    image = image.resize((350, 350), Image.ANTIALIAS)
    image = ImageTk.PhotoImage(image)
    root.imageLabel.config(image=image)
    root.imageLabel.photo = image
    button = Button(width=13, text="Registration Open", command=QROpen)
    button.grid(row=6, column=1, padx=5, pady=5, columnspan=1)
    buton = Button(width=13, text="Registration Closed", command=QRClose)
    buton.grid(row=6, column=3, padx=5, pady=5, columnspan=1)
    
def QRClear():        
    qrInput.set("")
    qrphno.set("")
    qrmail.set("")
    qrevent.set("")
    image = Image.open("./resc/done.png")
    image = image.resize((350, 350), Image.ANTIALIAS)
    image = ImageTk.PhotoImage(image)
    root.imageLabel.config(image=image)
    root.imageLabel.photo = image
    
def QROpen():
    print("OPEN")
    """
    reactivate all disabled fields & buttons
    """
    
def QRClose():
    print("Close")        
    """
    ask for CAPTCHA type verification, then disable 
    all fields and buttons except Registration Open,
    break databases according to event names
    """
    
def QRCodeGenerate():    
    qrString = qrInput.get()
    qrPhno = qrphno.get()
    qrMail = qrmail.get()
    qrEvent = qrevent.get()
    if (qrString != '') and (qrPhno != '') and (qrMail != '') and (qrEvent != ''):
        if (len(qrPhno) == 10):
            regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
            if(re.search(regex,qrMail)):  
                content = qrString + "-" + qrPhno
                qrGenerate = pyqrcode.create(content)
                qrCodePath = './data/'
                qrCodeName = qrCodePath + qrPhno + ".png"
                qrGenerate.png(qrCodeName, scale=10)
                image = Image.open(qrCodeName)
                image = image.resize((350, 350), Image.ANTIALIAS)
                image = ImageTk.PhotoImage(image)
                root.imageLabel.config(image=image)
                root.imageLabel.photo = image
                QRdatamgSQL()
                QRdatamgXL()
            else:  
                messagebox.showerror("ERROR", "Invalid Email ID")
        else:
            messagebox.showerror("ERROR", "Invalid Phone Number")
    else:
        messagebox.showerror("ERROR", "Fields Incomplete")

def QRdatamgXL():
    wb = load_workbook("./data/regdata.xlsx")
    sheet = wb.active
    row = ((qrInput.get(), qrphno.get(), qrmail.get(), qrevent.get()))
    sheet.append(row)
    wb.save("./data/regdata.xlsx")

def QRdatamgSQL():
    print("SHARAD")
    """
    ek new function bana to check if data 
    being entered is already in db.
    ek new function bana for SQL db creation,
    iss function se SQL db connection and
    add collected data to database.
    database ke table mei 5 columns rakh,
    4 for collected data, 1 for marking present.
    add if to check if email already in db,
    if yes give GUI prompt to confirms    
    """

root = tk.Tk()
root.title("QR Generator")        
#root.geometry("405x575")
root.geometry("600x600")
root.resizable(False, False)
root.config(background="green")
qrInput = StringVar()
qrphno = StringVar()
qrmail = StringVar()
workbook = Workbook()
qrevent = StringVar()
CreateWidgets()
root.mainloop()

